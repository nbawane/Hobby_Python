from openpyxl import load_workbook
wb = load_workbook(filename=r'C:\registervalidation\Product-Validation-Registers.xlsx', read_only=True)
ws = wb['CSD']
columndict = {}
dim = ws.calculate_dimension()
startrow = int(dim[1:2])

cardtype = 	'64GB SD Extreme'
print dim
for colindex,column in enumerate(range(ws.max_column)):
	columndict[colindex] = []

for rowindex,row in enumerate(ws.rows):
	if rowindex<startrow:
			continue
	for index,cell in enumerate(row):
		columndict[index].append(str(cell.value))

for key in columndict:

	if columndict[key][0] == cardtype:
		print columndict[key]