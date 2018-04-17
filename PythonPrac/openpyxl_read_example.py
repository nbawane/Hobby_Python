from openpyxl import load_workbook

wb = load_workbook(filename='C:\Prod_val\Product-Validation-Registers.xlsx',read_only=True)	#open the xlxs and load it to memory, callsed as workbook
ws = wb['SCR']	#load a particular worksheet
ws.iter_rows()

#load memory
memory_types = {}
data_start = 0

def get_memory_types():
	global data_start
	field_found = False
	for row in ws.rows:
		data_start += 1
		for cell in row:
			if 'Fields' == cell.value:
				field_found = True
				break
		for index,cell in enumerate(row):
			if cell.value:
				memory_types[cell.value]=index
		if field_found:
			break
	return memory_types


def get_memory_properties():
	# print(ws.max_row)
	maxrow = ws.max_row
	memory = get_memory_types()
	mem_prop = {}
	# print(ws.cell(row=data_start+1,column=(memory['400GB uSD Extreme'])+1).value)
	for rownum in range(maxrow):
		mem_prop[ws.cell(row=data_start + rownum, column=(memory['Fields']) + 1).value] = \
			(ws.cell(row=data_start + rownum, column=(memory['400GB uSD Extreme']) + 1).value)
	return mem_prop
print(get_memory_properties())



