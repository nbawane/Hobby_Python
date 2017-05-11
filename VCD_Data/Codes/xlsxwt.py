"""
********************************************************************************
* MODULE      : xlsxwt.py
* FUNCTION    : contains the function to write data to .xlsx file in predefined pattern
* PROGRAMMER  : Nitesh Bawane
* DATE(ORG)   : 01/03/2017
* REMARKS     : NA
* COPYRIGHT   : Copyright (C) 2015 SanDisk Corporation
*------------------------------------------------------------------------------*
* Revision History : 1.0
********************************************************************************
"""

from openpyxl import Workbook
import re

class write_to_excel:

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.location_x = 1
        self.location_y = 1

    def write_at_location(self,x,y,data):
        '''
        write at provided location in excel
        :param x: row location , starts from 1
        :param y: column location , starts from 1
        :param data: data needs to writen at the location
        :return: None
        '''
        self.location_x = x
        self.location_y = y
        if self.location_x<=0 or self.location_y<=0:
            raise Exception("location starts from 1, position provided (%s,%s)"%(self.location_x,self.location_y))
        else:
            self.ws.cell(row = self.location_x,column = self.location_y,value=data)

    def create_header(self):
        header = ['Sequence No.','Sequence description','Atomic Command Sequence']
        header_len = len(header)
        for column_index in range(1,header_len+1):
            self.write_at_location(1,column_index,header[column_index-1])

    def file_save(self,filename):
        """
        Must be called after all the alteration to the excel object has been done
        :param filename: give the appropriate filename without extension
        :return: None
        """
        seq = (filename,'.xlsx')
        filename = ''.join(seq)
        self.wb.save(filename = filename)
        print "result saved in filename : ",filename

    def read_from_location(self,x,y):
        pass

    def EraseCells(self,rows,columns):
        """
        function is used to clean the cells,writes None to location specified
        :param x: location of cell
        :param y: location of cell
        :return: None
        """

        for row in range(1,rows+1):
            for column in range(1,columns+1):
                print "Erasing (%s,%s)" % (row, column)
                self.ws.cell(row=row, column=column, value='')

    def Concat(self,loc_towrite,location_Row,location_Column):
        '''
        concatenate the rows of excel to one cell
        :param loc_towrite: cell location to store the the concatenated result
        :return:None
        '''
        arg = []
        joiner = ','
        for i in range(1,location_Row):
            str = 'A{}'.format(i)
            arg.append(str)
        #print arg
        arg = joiner.join(arg)
        formula = "=CONCATENATE(%s)"%arg
        self.ws[loc_towrite] = formula

    def writerange(self,start_corner,end_corner,data = None):
        '''
        to write a data on large range of block at time
        to be used to write single block also
        to erase the block data need not be passed
        not used as of now
        :param start_corner:
        :param end_corner:
        :param data:
        :return:None
        '''
        start_num = re.search('(?<=[A-Z])[0-9]*',start_corner)
        start_num = start_num.group()
        end_num = re.search('(?<=[A-Z])[0-9]*',end_corner)
        end_num = end_num.group()
        start_letter = re.search('[A-Z]',start_corner)
        start_letter = start_letter.group()
        end_letter = re.search('[A-Z]',end_corner)
        end_letter = end_letter.group()
        for j in range(ord(start_letter),ord(end_letter)+1):
            for i in range(int(start_num),int(end_num)+1):
                loc = '{}{}'.format(chr(j),i)
                self.ws[loc] = data



#
#
# if __name__ == "__main__":
#     obj = write_to_excel()
#     wb = Workbook()
#     ws = wb.active
#     obj.create_header(ws)
#     wb.save('demoxl--.xlsx')
